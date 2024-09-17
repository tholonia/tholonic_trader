"""
ExcelReporterClass.py: Excel Report Generation for Trading Bot

This module contains the ExcelReporter class, which is responsible for
creating and managing Excel reports for the trading bot's performance.

Key Features:
1. Creation and loading of Excel workbooks
2. Writing headers and appending data rows
3. Applying color coding to rows based on trading signals
4. Automatic column width adjustment for better readability

Main Class:
- ExcelReporter: Handles Excel file operations and report formatting

Dependencies:
- openpyxl: For Excel file manipulation
- os: For file system operations

Usage:
Instantiate the ExcelReporter class and use its methods to create
and populate Excel reports with trading data. This class is typically
used in conjunction with the main trading bot script to generate
performance reports.

Example:
    reporter = ExcelReporter("trading_report.xlsx")
    reporter.write_header(["Date", "Price", "Signal"])
    reporter.append_row(["2023-01-01", 50000, "BUY"])
    reporter.apply_row_color(2, "BUY")
    reporter.adjust_column_width()
    reporter.save()

Note: Ensure openpyxl is installed before using this class.
"""

import os
from openpyxl import Workbook, load_workbook
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import pandas as pd
from openpyxl.utils import get_column_letter, column_index_from_string

class ExcelReporter:
    def __init__(self, file_name="line_results.xlsx"):
        self.file_name = file_name
        self.wb = None
        self.ws = None
        self.load_or_create_workbook()



    def batch_write(self, all_rows, header, batch_size=20):
        """
        Write data to Excel in batches.

        :param all_rows: List of rows to write
        :param header: List of column headers
        :param batch_size: Number of rows to write in each batch (default 100)
        """
        if len(all_rows) >= batch_size:
            batch_df = pd.DataFrame(all_rows, columns=header)
            # print("------------------------------------------",batch_df)
            if os.path.exists(self.file_name):
                with pd.ExcelWriter(self.file_name, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
                    startrow = writer.sheets['Sheet1'].max_row
                    batch_df.to_excel(writer, index=False, header=False, startrow=startrow)
            else:
                with pd.ExcelWriter(self.file_name, mode='w', engine='openpyxl') as writer:
                    batch_df.to_excel(writer, index=False, header=True)
            return []  # Clear the batch
        return all_rows  # Return unchanged if batch size not reached

    def add_norm_close_column(self, testvars_report_filename):
            # Load the existing Excel file
            df = pd.read_excel(testvars_report_filename)

            # Identify the correct column names
            entry_price_col = 'entry_price'  # Update this if different
            cum_overhodl_col = 'cum_oh'  # Update this if different

            # Calculate 'norm_close'
            entry_price_min = df[entry_price_col].min()
            entry_price_max = df[entry_price_col].max()
            cum_overhodl_min = df[cum_overhodl_col].min()
            cum_overhodl_max = df[cum_overhodl_col].max()

            if entry_price_max != entry_price_min and cum_overhodl_max != cum_overhodl_min:
                df['norm_close'] = (
                    (df[entry_price_col] - entry_price_min) / (entry_price_max - entry_price_min)
                    * (cum_overhodl_max - cum_overhodl_min)
                    + cum_overhodl_min
                )
            else:
                df['norm_close'] = df[entry_price_col]

            # Load the existing workbook
            workbook = openpyxl.load_workbook(testvars_report_filename)

            # Get the active worksheet
            worksheet = workbook.active

            # Add the new column header
            new_col = worksheet.max_column + 1
            worksheet.cell(row=1, column=new_col, value='norm_close')

            # Add the 'norm_close' values
            for i, value in enumerate(df['norm_close'], start=2):
                worksheet.cell(row=i, column=new_col, value=value)

            # Format the new 'norm_close' column
            new_col_letter = get_column_letter(new_col)
            for cell in worksheet[new_col_letter]:
                cell.number_format = '0.00'

            # Adjust column width for the new column
            worksheet.column_dimensions[new_col_letter].width = 15

            # Save the changes
            workbook.save(testvars_report_filename)

            print(f"Updated {testvars_report_filename} with 'norm_close' column")

    def set_column_format(self, column, number_format):
        """
        Set the number format for an entire column.

        :param column: Column name
        :param number_format: Number format code (e.g., '0.00%', '$#,##0.00')
        """
        # print("Available columns:")
        # for idx, cell in enumerate(self.ws[1], start=1):
        #     print(f"- Column {get_column_letter(idx)}: {cell.value}")

        # Find the column letter by name
        for idx, cell in enumerate(self.ws[1], start=1):
            if cell.value and str(cell.value).strip().lower() == column.strip().lower():
                col_letter = get_column_letter(idx)
                break
        else:
            raise ValueError(f"Column '{column}' not found. Please check the column names printed above.")

        # Apply the number format to the entire column
        for row in range(2, self.ws.max_row + 1):  # Start from row 2 to skip header
            self.ws[f"{col_letter}{row}"].number_format = number_format

        # print(f"Format applied to column {col_letter} ({column})")


    def XXXset_column_format(self, column, number_format):
        """
        Set the number format for an entire column.

        :param column: Column name or letter
        :param number_format: Number format code (e.g., '0.00%', '$#,##0.00')
        """
        # # Print all column names for debugging
        # print("Available columns:")
        # for cell in self.ws[1]:
        #     print(f"- {cell.value}")

        if column.isalpha():
            col_letter = column.upper()
        else:
            # Find the column letter by name
            for idx, col in enumerate(self.ws[1], start=1):
                if col.value and col.value.strip().lower() == column.strip().lower():
                    col_letter = get_column_letter(idx)
                    break
            else:
                raise ValueError(f"Column '{column}' not found. Please check the column names printed above.")

        col = self.ws[col_letter]
        for cell in col[2:]:  # Skip the header row
            cell.number_format = number_format

    def xxset_column_format(self, column, number_format):
        """
        Set the number format for an entire column.

        :param column: Column letter (e.g., 'A', 'B') or number (1, 2, ...)
        :param number_format: Number format code (e.g., '0.00%', '$#,##0.00')
        """
        if isinstance(column, int):
            column = get_column_letter(column)

        for cell in self.ws[column]:
            cell.number_format = number_format
    def load_or_create_workbook(self):
        if os.path.exists(self.file_name):
            self.wb = load_workbook(self.file_name)
            self.ws = self.wb.active
        else:
            self.wb = Workbook()
            self.ws = self.wb.active

    def write_header(self, headers):
        if self.ws.max_row == 1:  # Only write header if sheet is empty
            self.ws.append(headers)

    def append_row(self, row_data):
        self.ws.append(row_data)
        return self.ws.max_row

    def apply_row_color(self, row_number, signal):
        red_fill = PatternFill(start_color="FFf5b7b1", end_color="FFf5b7b1", fill_type="solid")
        green_fill = PatternFill(start_color="FFd5f5e3", end_color="FFd5f5e3", fill_type="solid")

        fill = red_fill if signal == "BUY" else green_fill if signal == "SELL" else None

        if fill:
            for cell in self.ws[row_number]:
                cell.fill = fill

    def adjust_column_width(self):
        for column in self.ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            self.ws.column_dimensions[column_letter].width = adjusted_width

    def add_table(self):
        table_name = "TradingData"
        table_ref = f"A1:{get_column_letter(self.ws.max_column)}{self.ws.max_row}"

        # Check if the table already exists
        if table_name not in self.ws.tables:
            tab = Table(displayName=table_name, ref=table_ref)
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                   showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            tab.tableStyleInfo = style
            self.ws.add_table(tab)
        else:
            # Update existing table range
            self.ws.tables[table_name].ref = table_ref

    def save(self):
        self.wb.save(self.file_name)

    def update_report(self, data, headers):
        self.write_header(headers)
        row_number = self.append_row(data)
        self.apply_row_color(row_number, data[2])  # Assuming 'signal' is the 3rd item in data
        self.adjust_column_width()
        self.add_table()
        self.save()